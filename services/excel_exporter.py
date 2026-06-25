import os
from datetime import datetime
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, numbers
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.chart import BarChart, LineChart, Reference
import config
from database.db import get_db
import models.aggregates as agg
import models.net_worth as nw_model


def _df_transactions():
    from database.db import get_db
    rows = get_db().execute(
        "SELECT t.date, a.name as account, t.description, t.amount, t.category, t.category_source, t.notes "
        "FROM transactions t JOIN accounts a ON t.account_id=a.id ORDER BY t.date DESC"
    ).fetchall()
    return pd.DataFrame([dict(r) for r in rows])


def _df_net_worth():
    rows = get_db().execute(
        "SELECT snapshot_date, total_assets, total_liabilities, net_worth, notes "
        "FROM net_worth_snapshots ORDER BY snapshot_date ASC"
    ).fetchall()
    return pd.DataFrame([dict(r) for r in rows])


def _df_splits():
    rows = get_db().execute(
        """SELECT o.title, o.outing_date, op.name as person, op.is_paid,
                  SUM(oli.per_person_amount) as amount_owed
           FROM outing_participants op
           JOIN outings o ON o.id=op.outing_id
           JOIN outing_line_items oli ON oli.outing_id=o.id AND oli.paid_by_me=1
           GROUP BY op.id
           ORDER BY op.is_paid ASC, o.outing_date DESC"""
    ).fetchall()
    return pd.DataFrame([dict(r) for r in rows])


def _header_row(ws, headers, fill_color="1F4E79"):
    fill = PatternFill("solid", fgColor=fill_color)
    font = Font(color="FFFFFF", bold=True)
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center")


def _write_df(ws, df):
    _header_row(ws, list(df.columns))
    for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=False), 2):
        for c_idx, val in enumerate(row, 1):
            ws.cell(row=r_idx, column=c_idx, value=val)
    ws.auto_filter.ref = ws.dimensions


def export_excel():
    os.makedirs(config.OUTPUT_FOLDER, exist_ok=True)
    filename = f"finance_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    path = os.path.join(config.OUTPUT_FOLDER, filename)

    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # remove default sheet

    # --- Sheet 1: All Transactions ---
    df_tx = _df_transactions()
    ws1 = wb.create_sheet("Transactions")
    if not df_tx.empty:
        _write_df(ws1, df_tx)
        ws1.column_dimensions["C"].width = 40
        ws1.column_dimensions["A"].width = 12

    # --- Sheet 2: Spending by Category ---
    ws2 = wb.create_sheet("By Category")
    if not df_tx.empty:
        spend = df_tx[df_tx["amount"] < 0].copy()
        spend["amount"] = spend["amount"].abs()
        pivot = spend.groupby("category")["amount"].sum().sort_values(ascending=False).reset_index()
        pivot.columns = ["Category", "Total Spent"]
        _write_df(ws2, pivot)
        ws2.column_dimensions["A"].width = 20
        ws2.column_dimensions["B"].width = 15

        # Bar chart
        chart = BarChart()
        chart.title = "Spending by Category"
        chart.y_axis.title = "Amount ($)"
        data_ref  = Reference(ws2, min_col=2, min_row=1, max_row=len(pivot) + 1)
        cats_ref  = Reference(ws2, min_col=1, min_row=2, max_row=len(pivot) + 1)
        chart.add_data(data_ref, titles_from_data=True)
        chart.set_categories(cats_ref)
        chart.shape = 4
        ws2.add_chart(chart, "D2")

    # --- Sheet 3: Net Worth History ---
    ws3 = wb.create_sheet("Net Worth")
    df_nw = _df_net_worth()
    if not df_nw.empty:
        _write_df(ws3, df_nw)
        ws3.column_dimensions["A"].width = 14

        chart = LineChart()
        chart.title = "Net Worth Over Time"
        chart.y_axis.title = "Amount ($)"
        data_ref = Reference(ws3, min_col=4, min_row=1, max_row=len(df_nw) + 1)
        cats_ref = Reference(ws3, min_col=1, min_row=2, max_row=len(df_nw) + 1)
        chart.add_data(data_ref, titles_from_data=True)
        chart.set_categories(cats_ref)
        ws3.add_chart(chart, "F2")

    # --- Sheet 4: Outstanding Splits ---
    ws4 = wb.create_sheet("Splits Owed")
    df_sp = _df_splits()
    if not df_sp.empty:
        _write_df(ws4, df_sp)
        ws4.column_dimensions["A"].width = 25
        ws4.column_dimensions["C"].width = 18

    # --- Sheet 5: Cashflow pivot (sheet-parity, current year) ---
    year = datetime.now().year
    _write_cashflow_sheet(wb.create_sheet(f"Cashflow {year}"), agg.build_year_dashboard(year))

    # --- Sheet 6: Net worth by asset class (monthly series) ---
    _write_class_sheet(wb.create_sheet("NW by Class"))

    # --- Sheet 7: Account × month grid (the sheet's Total Assets view) ---
    _write_grid_sheet(wb.create_sheet(f"Assets {year}"), nw_model.get_year_grid(year))

    wb.save(path)
    return path


_BOLD = Font(bold=True)
_MONEY_FMT = "#,##0"


def _pivot_row(ws, label, values, bold=False):
    ws.append([label] + [round(v, 2) if v else None for v in values])
    r = ws.max_row
    if bold:
        ws.cell(row=r, column=1).font = _BOLD
    for c in range(2, len(values) + 2):
        cell = ws.cell(row=r, column=c)
        cell.number_format = _MONEY_FMT
        if bold:
            cell.font = _BOLD


def _write_cashflow_sheet(ws, d):
    """Income/Expenses/Net Income pivot exactly like /dashboard (and the old sheet)."""
    months = [m[:3] for m in d["months"]]
    _header_row(ws, [f"Cashflow {d['year']}"] + months + ["Year"])
    for r in d["income_rows"]:
        _pivot_row(ws, r["name"], r["by_month"] + [r["total"]])
    _pivot_row(ws, "Total Income", d["income_total"] + [d["ytd"]["income"]], bold=True)
    ws.append([])
    for r in d["expense_rows"]:
        _pivot_row(ws, r["name"], r["by_month"] + [r["total"]])
    _pivot_row(ws, "Total Expenses", d["expense_total"] + [d["ytd"]["expenses"]], bold=True)
    ws.append([])
    _pivot_row(ws, "Net Income", d["net"] + [d["ytd"]["net"]], bold=True)
    _pivot_row(ws, "Investments", d["investments"] + [d["ytd"]["investments"]])
    _pivot_row(ws, "FCF", d["fcf"] + [d["ytd"]["fcf"]], bold=True)
    _pivot_row(ws, "Cumulative Net Income", d["cumulative"] + [d["ytd"]["net"]])
    ws.column_dimensions["A"].width = 22


def _write_class_sheet(ws):
    dates, series = nw_model.get_class_series()
    _header_row(ws, ["Date"] + [c.title() for c in nw_model.CLASS_ORDER] + ["Net Worth"])
    for i, dt in enumerate(dates):
        vals = [series[c][i] for c in nw_model.CLASS_ORDER]
        ws.append([dt] + vals + [sum(vals)])
        for c in range(2, len(vals) + 3):
            ws.cell(row=ws.max_row, column=c).number_format = _MONEY_FMT
    ws.column_dimensions["A"].width = 12
    if dates:
        chart = LineChart()
        chart.title = "Net Worth by Asset Class"
        data_ref = Reference(ws, min_col=2, max_col=len(nw_model.CLASS_ORDER) + 2,
                             min_row=1, max_row=len(dates) + 1)
        cats_ref = Reference(ws, min_col=1, min_row=2, max_row=len(dates) + 1)
        chart.add_data(data_ref, titles_from_data=True)
        chart.set_categories(cats_ref)
        ws.add_chart(chart, "H2")


def _write_grid_sheet(ws, grid):
    months = grid["months"]
    _header_row(ws, ["Account"] + [f"{m:02d}" for m in months])
    for a in grid["accounts"]:
        ws.append([a["name"]] + [a["by_month"].get(m) for m in months])
        for c in range(2, len(months) + 2):
            ws.cell(row=ws.max_row, column=c).number_format = _MONEY_FMT
    ws.append([])
    for cls in nw_model.CLASS_ORDER:
        totals = grid["class_totals"].get(cls)
        if totals:
            _pivot_row(ws, cls.title(), [totals.get(m, 0) for m in months], bold=True)
    _pivot_row(ws, "Net Worth", [grid["net_worth"].get(m, 0) for m in months], bold=True)
    ws.column_dimensions["A"].width = 24
