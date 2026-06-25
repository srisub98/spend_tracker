#!/usr/bin/env python3
"""
One-time bootstrap of historical data from the Google Sheet workbook (docs/PRD.md Phase 1).

Input: data/bootstrap/expenses.xlsx (File > Download > Microsoft Excel of the sheet).

Imports:
  - "Budget YYYY" tabs -> monthly_summaries (income lines, expense categories,
    Investments row) and net_worth_snapshots + snapshot_account_balances
    (one snapshot per month-end, source='sheet'). Accounts are created on first
    sight with normalized names and asset classes.
  - "Transactions" tab (card-level rows he tracked from Apr 2025 onward) -> real
    transactions rows, import_batch_id='sheet-bootstrap'. These are for browsing/
    drill-down only: the Budget aggregates stay authoritative for months before
    LIVE_START_MONTH because the tab is card-only (no rent, no income).

Validation: recomputes each month's net worth from imported balances and compares
to the sheet's own Net Worth / Cumulative Assets row; same for Total Income /
Total Expenses vs imported line items. Discrepancies > $1 are printed.

Usage:
  python3 scripts/bootstrap_from_sheet.py            # refuses if sheet data already present
  python3 scripts/bootstrap_from_sheet.py --reset    # wipe prior sheet-sourced data, reimport
"""

import argparse
import calendar
import datetime
import json
import os
import sqlite3
import sys
import warnings

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

import openpyxl  # noqa: E402
import config  # noqa: E402
from database.db import _migrate, _seed  # noqa: E402

warnings.filterwarnings("ignore")  # openpyxl: unsupported sparklines etc.

MONTHS = list(calendar.month_name)[1:]  # January..December
TX_BATCH_ID = "sheet-bootstrap"

# Sheet row label -> canonical account (name, type, institution, asset_class).
# TDAmeritrade became Schwab when TDA was acquired — one continuous account.
ACCOUNT_MAP = {
    "CitiBank":              ("CitiBank Checking",  "checking",  "Citi",       "cash"),
    "CapitalOne":            ("CapitalOne Savings", "savings",   "CapitalOne", "cash"),
    "CapitalOne Savings":    ("CapitalOne Savings", "savings",   "CapitalOne", "cash"),
    "CaptialOne Checking":   ("CapitalOne Checking","checking",  "CapitalOne", "cash"),
    "CDs":                   ("CDs",                "savings",   "Citi",       "cash"),
    "CitiBank CDs":          ("CDs",                "savings",   "Citi",       "cash"),
    "TDAmeritrade":          ("Schwab - personal",  "brokerage", "Schwab",     "stocks"),
    "TDAmeritrade / Schwab": ("Schwab - personal",  "brokerage", "Schwab",     "stocks"),
    "Schwab - personal":     ("Schwab - personal",  "brokerage", "Schwab",     "stocks"),
    "Schwab - Meta":         ("Schwab - Meta",      "brokerage", "Schwab",     "stocks"),
    "401K":                  ("401K",               "brokerage", None,         "retirement"),
    "Etrade":                ("Etrade",             "brokerage", "Etrade",     "stocks"),
    "Etrade Balance":        ("Etrade",             "brokerage", "Etrade",     "stocks"),
    "Vanguard":              ("Vanguard",           "brokerage", "Vanguard",   "stocks"),
    "Fundrise":              ("Fundrise",           "other",     "Fundrise",   "other"),
    "Venmo/AppleCash":       ("Venmo/AppleCash",    "other",     None,         "cash"),
    "Coinbase":              ("Coinbase",           "other",     "Coinbase",   "other"),
}
# Rows inside the Total Assets block that are not accounts (validation/rollup/junk).
NON_ACCOUNT_ROWS = {"Net Worth", "Cumulative Assets", "Stocks", "Cash", "Retirement",
                    "Other", "Unvested RSUs"}

INCOME_NORMALIZE = [          # (substring of row label, canonical income line)
    ("rsu", "RSU Vest"),
    ("interest", "Interest"),
    ("paycheck", "Paycheck"),
    ("other income", "Other Income"),
]

EXPENSE_NORMALIZE = {
    "Misc.": "Misc",
    "Rent": "Rent + Utilities",
    "Utilities": "Rent + Utilities",
    "Rent & Utilities": "Rent + Utilities",
    "Running /Cycling/Fitness": "Fitness",
}


def num(v):
    """Cell value -> float, or None for blanks/labels/junk (e.g. '119.489.81')."""
    if isinstance(v, (int, float)):
        return float(v)
    if isinstance(v, str):
        s = v.strip().replace("$", "").replace(",", "")
        try:
            return float(s)
        except ValueError:
            return None
    return None


def label(ws, r, c):
    v = ws.cell(r, c).value
    return str(v).strip() if isinstance(v, str) else ""


def month_end(year, month_idx):
    return f"{year}-{month_idx:02d}-{calendar.monthrange(year, month_idx)[1]}"


def month_cols(ws, row):
    """Map month index (1-12) -> column, from a header row containing January..December."""
    cols = {}
    for c in range(1, ws.max_column + 1):
        v = label(ws, row, c)
        if v in MONTHS:
            cols[MONTHS.index(v) + 1] = c
    return cols


def parse_budget_tab(ws, year):
    """Returns dict with income/expenses/investments (per-month, normalized+merged),
    balances per account, and the sheet's own totals for validation."""
    out = {
        "income": {},        # (month_idx, line) -> amount
        "expenses": {},      # (month_idx, category) -> amount
        "investments": {},   # month_idx -> amount
        "balances": {},      # (month_idx, sheet_label) -> balance
        "sheet_income_total": {}, "sheet_expense_total": {}, "sheet_net_worth": {},
    }
    section, cols = None, {}
    for r in range(1, ws.max_row + 1):
        a, b = label(ws, r, 1), label(ws, r, 2)

        if a in ("Income", "Expenses", "Net Income"):
            section, cols = a, month_cols(ws, r)
            continue
        if b == "Total Assets":
            section, cols = "Total Assets", month_cols(ws, r)
            continue
        if not section or not b or not cols:
            continue

        values = {m: num(ws.cell(r, c).value) for m, c in cols.items()}

        if section == "Income":
            if b == "Total Income":
                out["sheet_income_total"] = {m: v for m, v in values.items() if v is not None}
                continue
            line = next((canon for frag, canon in INCOME_NORMALIZE if frag in b.lower()), None)
            if line is None:
                continue
            for m, v in values.items():
                if v:
                    out["income"][(m, line)] = out["income"].get((m, line), 0.0) + v

        elif section == "Expenses":
            if b.startswith("Total Expense"):
                out["sheet_expense_total"] = {m: v for m, v in values.items() if v is not None}
                continue
            cat = EXPENSE_NORMALIZE.get(b, b)
            for m, v in values.items():
                if v:
                    out["expenses"][(m, cat)] = out["expenses"].get((m, cat), 0.0) + v

        elif section == "Net Income":
            if b == "Investments":
                for m, v in values.items():
                    if v:
                        out["investments"][m] = v

        elif section == "Total Assets":
            if b in ("Net Worth", "Cumulative Assets"):
                out["sheet_net_worth"] = {m: v for m, v in values.items() if v is not None}
                continue
            if b in NON_ACCOUNT_ROWS:
                continue
            if b not in ACCOUNT_MAP:
                print(f"  WARNING {year}: unknown Total Assets row {b!r} — skipped")
                continue
            for m, v in values.items():
                if v is not None:
                    out["balances"][(m, b)] = v
    return out


def parse_transactions_tab(ws):
    """Card-level rows: Date, Purchase, Price, Category, Split, Notes, Final Cost,
    Amount Owed, Source, Month, Year. amount = -Final Cost (his share after splits)."""
    rows = []
    for r in range(2, ws.max_row + 1):
        d = ws.cell(r, 1).value
        if not isinstance(d, datetime.datetime):
            continue
        desc = str(ws.cell(r, 2).value or "").strip()
        cost = num(ws.cell(r, 7).value)
        if cost is None:
            cost = num(ws.cell(r, 3).value)
        if not desc or cost is None:
            continue
        cat = str(ws.cell(r, 4).value or "").strip()
        cat = EXPENSE_NORMALIZE.get(cat, cat)
        if cat in ("", "TODO"):
            cat = None
        split = str(ws.cell(r, 5).value or "").strip()
        notes = str(ws.cell(r, 6).value or "").strip()
        if split:
            price, owed = num(ws.cell(r, 3).value), num(ws.cell(r, 8).value) or 0.0
            notes = (notes + " " if notes else "") + \
                f"[split with {split}: total ${price:,.2f}, my share ${cost:,.2f}, owed ${owed:,.2f}]"
        rows.append({
            "date": d.strftime("%Y-%m-%d"),
            "description": desc,
            "amount": -cost,
            "category": cat,
            "category_source": "user" if cat else None,
            "notes": notes or None,
            "raw": json.dumps({"source": str(ws.cell(r, 9).value or ""), "price": ws.cell(r, 3).value}),
        })
    return rows


def ensure_account(db, sheet_label, cache):
    name, type_, institution, asset_class = ACCOUNT_MAP[sheet_label]
    if name in cache:
        return cache[name]
    row = db.execute("SELECT id FROM accounts WHERE name=?", (name,)).fetchone()
    if row:
        cache[name] = row[0]
    else:
        cur = db.execute(
            "INSERT INTO accounts (name, type, institution, asset_class) VALUES (?,?,?,?)",
            (name, type_, institution, asset_class))
        cache[name] = cur.lastrowid
    return cache[name]


def main():
    ap = argparse.ArgumentParser(description=__doc__)
    ap.add_argument("--file", default="data/bootstrap/expenses.xlsx")
    ap.add_argument("--reset", action="store_true",
                    help="delete previously bootstrapped sheet data before importing")
    args = ap.parse_args()

    db = sqlite3.connect(config.DB_PATH)
    db.row_factory = sqlite3.Row
    with open(os.path.join("database", "schema.sql")) as f:
        db.executescript(f.read())
    _migrate(db)
    _seed(db)

    existing = db.execute("SELECT COUNT(*) FROM monthly_summaries WHERE source='sheet'").fetchone()[0]
    if existing and not args.reset:
        sys.exit(f"Sheet data already bootstrapped ({existing} summary rows). Re-run with --reset.")
    if args.reset:
        db.execute("DELETE FROM monthly_summaries WHERE source='sheet'")
        db.execute("DELETE FROM snapshot_account_balances WHERE snapshot_id IN "
                   "(SELECT id FROM net_worth_snapshots WHERE source='sheet')")
        db.execute("DELETE FROM holdings WHERE snapshot_id IN "
                   "(SELECT id FROM net_worth_snapshots WHERE source='sheet')")
        db.execute("DELETE FROM net_worth_snapshots WHERE source='sheet'")
        db.execute("DELETE FROM transactions WHERE import_batch_id=?", (TX_BATCH_ID,))

    wb = openpyxl.load_workbook(args.file, data_only=True)
    acct_cache = {}
    n_summaries = n_snapshots = n_warn = 0

    for ws in wb.worksheets:
        if not ws.title.startswith("Budget "):
            continue
        year = int(ws.title.split()[-1])
        data = parse_budget_tab(ws, year)
        print(f"{ws.title}: {len(data['income'])} income cells, {len(data['expenses'])} expense cells, "
              f"{len(data['investments'])} investment months, {len(data['balances'])} balance cells")

        for (m, line), v in data["income"].items():
            db.execute("INSERT OR REPLACE INTO monthly_summaries (month, category, kind, amount, source) "
                       "VALUES (?,?,?,?, 'sheet')", (f"{year}-{m:02d}", line, "income", round(v, 2)))
            n_summaries += 1
        for (m, cat), v in data["expenses"].items():
            db.execute("INSERT OR REPLACE INTO monthly_summaries (month, category, kind, amount, source) "
                       "VALUES (?,?,?,?, 'sheet')", (f"{year}-{m:02d}", cat, "expense", round(v, 2)))
            n_summaries += 1
        for m, v in data["investments"].items():
            db.execute("INSERT OR REPLACE INTO monthly_summaries (month, category, kind, amount, source) "
                       "VALUES (?,?,?,?, 'sheet')", (f"{year}-{m:02d}", "Investments", "investment", round(v, 2)))
            n_summaries += 1

        # Snapshots: one per month that has any nonzero balance
        by_month = {}
        for (m, sheet_label), v in data["balances"].items():
            by_month.setdefault(m, []).append((sheet_label, v))
        for m in sorted(by_month):
            balances = by_month[m]
            total = sum(v for _, v in balances)
            if total == 0:
                continue
            cur = db.execute(
                "INSERT INTO net_worth_snapshots (snapshot_date, total_assets, total_liabilities, "
                "net_worth, notes, source) VALUES (?,?,0,?,?, 'sheet')",
                (month_end(year, m), round(total, 2), round(total, 2), f"bootstrap from {ws.title}"))
            for sheet_label, v in balances:
                acct_id = ensure_account(db, sheet_label, acct_cache)
                db.execute("INSERT INTO snapshot_account_balances (snapshot_id, account_id, balance) "
                           "VALUES (?,?,?)", (cur.lastrowid, acct_id, v))
            n_snapshots += 1

            sheet_nw = data["sheet_net_worth"].get(m)
            if sheet_nw is not None and abs(sheet_nw - total) > 1.0:
                print(f"  MISMATCH {year}-{m:02d}: recomputed NW {total:,.2f} vs sheet {sheet_nw:,.2f} "
                      f"(diff {total - sheet_nw:+,.2f})")
                n_warn += 1

        # Validate income/expense line items vs the sheet's own totals
        for kind_label, items, sheet_totals in (
                ("income", data["income"], data["sheet_income_total"]),
                ("expenses", data["expenses"], data["sheet_expense_total"])):
            for m, sheet_v in sheet_totals.items():
                mine = sum(v for (mm, _), v in items.items() if mm == m)
                if abs(mine - sheet_v) > 1.0:
                    print(f"  MISMATCH {year}-{m:02d} {kind_label}: imported {mine:,.2f} "
                          f"vs sheet total {sheet_v:,.2f}")
                    n_warn += 1

    # Transactions tab -> real transactions under a dedicated import account
    tx_rows = parse_transactions_tab(wb["Transactions"]) if "Transactions" in wb.sheetnames else []
    n_tx = n_dup = 0
    if tx_rows:
        row = db.execute("SELECT id FROM accounts WHERE name=?", ("Card (sheet import)",)).fetchone()
        card_id = row[0] if row else db.execute(
            "INSERT INTO accounts (name, type, institution, is_liability) "
            "VALUES ('Card (sheet import)', 'credit', NULL, 1)").lastrowid
        for t in tx_rows:
            cur = db.execute(
                "INSERT OR IGNORE INTO transactions (account_id, date, description, amount, category, "
                "category_source, notes, raw_csv_row, import_batch_id) VALUES (?,?,?,?,?,?,?,?,?)",
                (card_id, t["date"], t["description"], t["amount"], t["category"],
                 t["category_source"], t["notes"], t["raw"], TX_BATCH_ID))
            n_tx += cur.rowcount
            n_dup += 1 - cur.rowcount

    db.commit()
    dates = db.execute("SELECT MIN(snapshot_date), MAX(snapshot_date) FROM net_worth_snapshots "
                       "WHERE source='sheet'").fetchone()
    print(f"\nDone: {n_summaries} summary rows, {n_snapshots} snapshots ({dates[0]} → {dates[1]}), "
          f"{n_tx} transactions imported ({n_dup} duplicate rows collapsed), "
          f"{len(acct_cache)} accounts, {n_warn} validation warnings.")
    print(f"Dashboards read sheet data for months < LIVE_START_MONTH ({config.LIVE_START_MONTH}).")


if __name__ == "__main__":
    main()
