"""Report generation against the synthetic demo dataset.

These ride on the `demo_*` fixtures (tests/python/conftest.py), which seed months
of sheet history, live transactions, net-worth snapshots + holdings, and bill
splits — enough that the Excel workbook and HTML dashboard render with real
content instead of empty stubs.
"""
import openpyxl
import pytest

import config


@pytest.fixture()
def output_dir(tmp_path, monkeypatch):
    """Redirect exports at a scratch folder so tests never touch data/exports/."""
    out = tmp_path / "exports"
    monkeypatch.setattr(config, "OUTPUT_FOLDER", str(out))
    return out


def test_demo_dataset_is_populated(demo_app):
    """The seeder lands the expected shape: categorized live transactions (incl. a
    couple awaiting review), sheet history across years, snapshots, and outings."""
    from database.db import get_db

    with demo_app.app_context():
        db = get_db()

        def scalar(sql):
            return db.execute(sql).fetchone()[0]

        assert scalar("SELECT COUNT(*) FROM transactions") == 23
        # Every live row is categorized; only the two Claude rows await review.
        assert scalar("SELECT COUNT(*) FROM transactions WHERE category IS NULL") == 0
        assert scalar("SELECT COUNT(*) FROM transactions WHERE category_source='claude'") == 2
        # Sheet history spans 2025 (full) + 2026 (pre-cutover) and never collides
        # with the live month.
        assert scalar("SELECT COUNT(*) FROM monthly_summaries WHERE month >= '2026-06'") == 0
        assert scalar("SELECT COUNT(*) FROM net_worth_snapshots") == 6
        assert scalar("SELECT COUNT(*) FROM holdings") == 3
        assert scalar("SELECT COUNT(*) FROM outings") == 3
        # Retirement + liabilities are represented for net-worth reports.
        assert scalar("SELECT COUNT(*) FROM accounts WHERE asset_class='retirement'") == 1
        assert scalar("SELECT COUNT(*) FROM accounts WHERE is_liability=1") == 2


def test_dashboard_spans_years_with_real_figures(demo_app):
    import models.aggregates as agg

    with demo_app.app_context():
        assert agg.available_years() == [2025, 2026]
        d = agg.build_year_dashboard(2026)
        # Income/expense/investment all flow through (sheet history + live June).
        assert d["ytd"]["income"] > 0
        assert d["ytd"]["expenses"] > 0
        assert d["ytd"]["investments"] > 0
        assert 0 < d["ytd"]["savings_rate"] < 1
        assert {"Paycheck", "RSU Vest", "Interest"} <= {r["name"] for r in d["income_rows"]}


def test_net_worth_class_series_includes_reclassified_cash(demo_app):
    import models.net_worth as nw

    with demo_app.app_context():
        dates, series = nw.get_class_series()
        assert len(dates) == 6
        assert all(v > 0 for v in series["stocks"])
        assert all(v > 0 for v in series["retirement"])
        # The June Schwab cash sweep ($4,386.47) is reclassified out of stocks
        # into cash, so the last cash point jumps above the prior month's.
        assert series["cash"][-1] > series["cash"][-2]


def test_excel_export_has_populated_sheets(demo_app, output_dir):
    from services.excel_exporter import export_excel

    with demo_app.app_context():
        path = export_excel()

    wb = openpyxl.load_workbook(path)
    assert wb.sheetnames == [
        "Transactions", "By Category", "Net Worth", "Splits Owed",
        "Cashflow 2026", "NW by Class", "Assets 2026",
    ]
    # Each data sheet has a header row plus real rows (not just the stub header).
    assert wb["Transactions"].max_row >= 20
    assert wb["By Category"].max_row > 1
    assert wb["Net Worth"].max_row == 7          # 6 snapshots + header
    assert wb["Splits Owed"].max_row > 1         # outstanding balances exist


def test_html_export_is_self_contained(demo_app, output_dir):
    from services.html_exporter import export_html

    with demo_app.app_context():
        path = export_html()

    html = open(path, encoding="utf-8").read()
    # Self-contained: Chart.js is inlined, no external script/CDN reference.
    assert "Chart" in html
    assert "cdn" not in html.lower()
    # Real content made it into the embedded JSON.
    assert "Net Worth" in html
    assert "Tahoe Ski Trip" in html or "Alex" in html
